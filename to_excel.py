
# to_excel.py

# In this file, we import the data, apply modifications to it and return a downloadable memory object
# that will be downloaded interactively.


import base64
import locale
from io import BytesIO
import openpyxl
from openpyxl.styles import Alignment, Font, Side, Border
from openpyxl.styles.colors import Color
from openpyxl.styles.fills import PatternFill
from pandas import ExcelWriter
import pandas as pd

all_rows = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
            'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE']

def set_border(ws, cell_range):
    rows = list(ws[cell_range])
    side = Side(border_style='thin', color="FF000000")
    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom)
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border


def get_borders(sheet, len_df):
    for row in all_rows:
        elem=f'{row}12:{row}{len_df}'
        set_border(sheet, elem)
    return sheet

def get_downlable_excel(df, len_df, filename='DASHBOARD Excel', text='Descarga como archivo Excel>>', title='Descargar DASHBOARD como excel',
                               color='#262730', color_fondo='#f0f2f6', text_decoration=None,):

    #locale.setlocale(locale.LC_ALL, '')
    xlsx_io = BytesIO()
    writer = ExcelWriter(xlsx_io, engine='openpyxl')

    df.to_excel(writer, sheet_name='Dataframe after NLP',startcol = 1, startrow=10, index=True, header = True)

    initial_sheet = writer.sheets['Dataframe after NLP']
    initial_sheet.sheet_view.showGridLines = False
    initial_sheet.row_dimensions[1].height = 37.5
    initial_sheet = get_borders(initial_sheet, len_df)

    initial_sheet.sheet_view.zoomScale = 110
    for col_ in ['F', 'J', 'Y', 'Z', 'AA', 'AB', 'AC']:
        initial_sheet.column_dimensions[col_].width = 80

    for col_ in ['A', 'B', 'C', 'D', 'E', 'G', 'H', 'I', 'K', 'L', 'M', 'N', 'O',
                 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'AD', 'AE']:
        initial_sheet.column_dimensions[col_].width = 25

    for row in all_rows:
        for col in range(1, 200):
            initial_sheet[f'{row}{col}'].alignment = Alignment(horizontal='center', vertical='center')

    # color titles
    for row in initial_sheet['C11:AE11']:
        for cell in row:
            initial_sheet[cell.coordinate].font = Font(name="Calibri", sz=14, bold=True)
            cell.fill = PatternFill(patternType='solid', fgColor=Color(rgb='B22222'))


    initial_sheet['B3'] = 'DYNAMIC DASHBOARD. REDDIT CITIES'
    initial_sheet['B3'].font = Font(name='Calibri', bold=True, size=20, color = '8ea9db')
    initial_sheet['B3'].alignment = Alignment(horizontal='left', vertical='center')
    #initial_sheet['D19'].value = ""

    writer.save()
    xlsx_io.seek(0)
    media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    data = base64.b64encode(xlsx_io.read()).decode("utf-8")
    href_data_downloadable = f'data:{media_type};base64,{data}'
    filename = f'{filename}.xlsx'
    if not text_decoration:
        text_decoration = 'none'
    href = f'<p style="background-color: {color_fondo}"><a href=' \
           f'"{href_data_downloadable}" download="{filename}" ' \
           f'style="text-decoration: {text_decoration}; ' \
           f'color: {color}" ' \
           f'title="{title}">{text}</a><p>'


    return href
